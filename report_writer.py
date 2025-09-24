# report_writer.py

"""
MODULE NAME: report_writer.py
PART OF: City Weather Report
AUTHOR: Joshua C. Stroud
VERSION: 3.0
INTERPRETER: Python 3.12.10
PURPOSE: Creates and formats Excel workbook with multiple worksheets
DEPENDENCIES: openpyxl, logging, config
LICENSE: GNU GPL v3.0

MODULE CHANGE LOG:

Module Version  |   Date        |   Description
---------------------------------------------------------------------------------------------------------
1.0             |   2025-09-24  |   Initial Release
---------------------------------------------------------------------------------------------------------
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
import logging
from config import headers

def format_worksheet(ws):
    header_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
    for col_num in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = Font(bold=True)
        cell.fill = header_fill
    for col in ws.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_length + 2

def create_excel_report(data, filename):
    wb = Workbook()

    # Sheet 1: City Weather Report
    ws1 = wb.active
    ws1.title = "City Weather Report"
    for col_num, header in enumerate(headers, start=1):
        ws1.cell(row=1, column=col_num, value=header.capitalize())
    for row_num, record in enumerate(data.values(), start=2):
        if not isinstance(record, dict):
            logging.warning(f"Invalid record at row {row_num}: {type(record)}")
            continue
        for col_num, key in enumerate(headers, start=1):
            ws1.cell(row=row_num, column=col_num, value=record.get(key))
    format_worksheet(ws1)

    # Sheet 2: Cities by Region
    ws2 = wb.create_sheet(title="Cities by Region")
    for col_num, header in enumerate(headers, start=1):
        ws2.cell(row=1, column=col_num, value=header.capitalize())
    sorted_records = sorted(data.values(), key=lambda x: x.get("region", "Unknown"))
    for row_num, record in enumerate(sorted_records, start=2):
        for col_num, key in enumerate(headers, start=1):
            ws2.cell(row=row_num, column=col_num, value=record.get(key))
    format_worksheet(ws2)

    wb.save(filename)
    logging.info(f"Workbook saved: {filename}")